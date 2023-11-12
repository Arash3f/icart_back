from random import randint
from uuid import UUID

import openpyxl
from sqlalchemy.ext.asyncio import AsyncSession

from src.cash.models import Cash
from src.core.security import hash_password
from src.credit.models import Credit
from src.location.crud import location as location_crud
from src.user.crud import user as user_crud
from src.organization.crud import organization as organization_crud
from src.user.models import User
from src.wallet.models import Wallet

dataframe = openpyxl.load_workbook("src/utils/icart-register.xlsx")

dataframe1 = dataframe.active


async def read_excel_file(db: AsyncSession, user_id: UUID):
    organization_user = await organization_crud.find_by_user_id(
        db=db,
        user_id="61675eef-458b-463d-89ee-e40c2cbe6f93",
    )

    error_message: list[str] = []

    # ? Calculate all rows
    row_count = 0
    finish = False
    row = 2
    while not finish:
        for col in dataframe1.iter_cols(1, 1):
            if col[row].value:
                row_count += 1
            else:
                finish = True
        row += 1

    final_generate_user_list = []
    final_append_user_list = []
    for row in dataframe1.iter_rows(3, row_count + 2):
        username = str(row[4].value)
        # ! Verify username len
        if len(username) != 11:
            text = "کاربر با شماره {} به دلیل نامعتبر بودن شماره همراه ({}) ثبت نشد! \n ".format(
                row[0].value,
                username,
            )
            error_message.append(text)

        else:
            # * find location
            location = await location_crud.find_by_name(db=db, name=str(row[8].value))
            # ! Verify location
            if not location:
                text = "کاربر با شماره {} به دلیل پیدا نشدن منطقه وارد شده ({}) ثبت نشد! \n ".format(
                    row[0].value,
                    row[8].value,
                )
                error_message.append(text)

            else:
                # * Check user not exist
                exist_user = await user_crud.check_by_username_and_national_code(
                    db=db,
                    username=str(row[4].value),
                    national_code=str(row[3].value),
                )
                # ! Verify user existence
                if exist_user:
                    if exist_user.credit.considered:
                        organization_user.total_considered_credit -= (
                            exist_user.credit.considered
                        )
                    organization_user.total_considered_credit += int(row[15].value)
                    db.add(organization_user)

                    # ? Append new User
                    exist_user.organization_id = organization_user.id
                    exist_user.credit.considered = int(row[15].value)
                else:
                    # * Update Organization considered credit
                    organization_user.total_considered_credit += int(row[15].value)
                    db.add(organization_user)

                    find_user = await user_crud.check_by_username_or_national_code(
                        db=db,
                        username=str(row[4].value),
                        national_code=str(row[3].value),
                    )
                    if find_user:
                        text = "کاربر با شماره {} به دلیل تکراری بودن شماره همراه و یا کد ملی ثبت نشد! \n ".format(
                            row[0].value,
                            row[8].value,
                        )
                        error_message.append(text)

                    else:
                        # ? Generate new User -> validation = False
                        new_user = User(
                            organization_id=organization_user.id,
                            first_name=str(row[1].value),
                            last_name=str(row[2].value),
                            national_code=str(row[3].value),
                            phone_number=str(row[4].value),
                            father_name=str(row[5].value),
                            birth_place=str(row[6].value),
                            location_id=location.id,
                            postal_code=str(row[9].value),
                            tel=str(row[10].value),
                            address=str(row[11].value),
                            personnel_number=str(row[12].value),
                            organizational_section=str(row[13].value),
                            job_class=str(row[14].value),
                            password=hash_password(str(123456789)),
                        )

                        final_generate_user_list.append(
                            new_user,
                        )

                        # # ? Create Credit
                        # credit = Credit(
                        #     user=new_user,
                        #     considered=int(row[15].value),
                        # )
                        #
                        # # ? Create Cash
                        # cash = Cash(
                        #     user=new_user,
                        # )
                        #
                        # # ? Create Wallet
                        # wallet_number = randint(100000, 999999)
                        # wallet = Wallet(
                        #     user=new_user,
                        #     number=wallet_number,
                        # )
                        #
                        # db.add(new_user)
                        # db.add(credit)
                        # db.add(cash)
                        # db.add(wallet)
                        #
                        # await db.commit()
                        # result_message.append(
                        #     "کاربر با کد ملی {} با موفقیت به سازمان شما پیوست".format(
                        #         str(row[3].value),
                        #     ),
                        # )

    print(error_message)
